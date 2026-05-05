# -*- coding: utf-8 -*-
"""2026年温州市招聘岗位信息汇总 - 详细版"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ===== Styles =====
DARK_BLUE = '1F4E79'
MEDIUM_BLUE = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
GREEN = '548235'
LIGHT_GREEN = 'E2EFDA'
RED = 'FF0000'
ORANGE = 'ED7D31'
LIGHT_ORANGE = 'FFF2CC'
GRAY = '808080'

HEADER_FILL = PatternFill('solid', fgColor=DARK_BLUE)
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
SUB_FILL = PatternFill('solid', fgColor=MEDIUM_BLUE)
SUB_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
CAT_FILL = PatternFill('solid', fgColor=LIGHT_BLUE)
CAT_FONT = Font(name='微软雅黑', bold=True, size=10)
NORMAL_FONT = Font(name='微软雅黑', size=9)
BOLD_FONT = Font(name='微软雅黑', bold=True, size=9)
RED_BOLD = Font(name='微软雅黑', bold=True, size=9, color=RED)
BORDER = Border(
    left=Side('thin', color='B0B0B0'),
    right=Side('thin', color='B0B0B0'),
    top=Side('thin', color='B0B0B0'),
    bottom=Side('thin', color='B0B0B0')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_sub(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUB_FILL
        cell.font = SUB_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_cat(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = CAT_FILL
        cell.font = CAT_FONT
        cell.alignment = LEFT
        cell.border = BORDER

def style_normal(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_left(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.border = BORDER

def add_title(ws, title, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE)
    ws.cell(row=row, column=1).alignment = CENTER

# ====================================================================
# Sheet 1: 总览
# ====================================================================
ws1 = wb.active
ws1.title = '招聘总览'
add_title(ws1, '2026年温州市公务员及编制类招聘信息汇总（详细版）', 8)

headers1 = ['序号', '招聘类型', '招聘范围', '职位/岗位数', '招聘人数', '报名时间', '笔试/面试时间', '备注/来源']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 8)

data1 = [
    ['1', '国考', '温州地区', '86个职位', '138人', '2025.10.15-10.24', '2025.11.30笔试', '连续3年减少，税务为主力'],
    ['2', '省考（四级联考）', '温州地区', '495个职位', '660人', '2025.11.6-11.11', '2025.12.7笔试', '较去年减379人，降幅超1/3'],
    ['3', '省考（山区海岛县专项）', '洞头/文成/泰顺', '-', '28人', '2026.4.8-4.12', '2026.4.26面试', '需已参加省考笔试且达线'],
    ['4', '事业单位（市级）', '温州市级', '-', '36名', '2026.3.23-3.27', '2026.4.25笔试', '综合应用能力+职业能力倾向'],
    ['5', '事业单位（鹿城区）', '鹿城区', '-', '21名', '2026.3.17-3.24', '2026.4.25笔试', ''],
    ['6', '事业单位（龙湾区）', '龙湾区', '-', '28名', '2026.3.18-3.25', '2026.4.25笔试', ''],
    ['7', '事业单位（瓯海区）', '瓯海区', '-', '28名', '2026.3.23-3.27', '2026.4.25笔试', ''],
    ['8', '事业单位（洞头区）', '洞头区', '-', '44名', '2026.3.19-3.25', '2026.4.25笔试', ''],
    ['9', '事业单位（龙港市）', '龙港市', '-', '23名', '2026.3.17-3.24', '2026.4.25笔试', ''],
    ['10', '事业单位（苍南县）', '苍南县', '-', '41名', '2026.3.17-3.23', '2026.4.25笔试', ''],
    ['11', '事业单位（文成县）', '文成县', '-', '63名', '2026.3.19-3.25', '2026.4.25笔试', ''],
    ['12', '国企（永嘉县）', '永嘉县', '-', '12名', '2026年3月', '2026年4月', '5家国企劳动合同制'],
    ['13', '国企（乐清市）', '乐清市', '-', '43名', '2026.3.18-3.20', '2026.4.11', '市属国企，最低3年'],
    ['14', '国企（洞头区）', '洞头区', '-', '10名', '2026.4.7-4.9', '-', '4家国企劳动合同制'],
    ['15', '硕博人才招引', '市级+知名企业', '65家单位', '445人', '2026.4.13-4.28', '待定', '事业编213+企业232，仅限硕博'],
]

for i, row in enumerate(data1, 4):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
    if i in [4, 5, 6]:  # 国考/省考/专项
        style_cat(ws1, i, 8)
    elif i in [7, 8, 9, 10, 11, 12, 13, 14]:  # 事业编
        style_normal(ws1, i, 8)
    elif i in [15, 16, 17]:  # 国企
        style_cat(ws1, i, 8)
    else:
        style_normal(ws1, i, 8)

# 合计行
r = 19
ws1.cell(row=r, column=1, value='合计')
ws1.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=RED)
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
ws1.cell(row=r, column=2, value='公务员: 826人(含专项28人) + 事业编: 284人 + 国企: 65人 + 硕博: 445人 = 总计 1620人')
ws1.cell(row=r, column=2).font = Font(name='微软雅黑', bold=True, size=11, color=RED)
ws1.cell(row=r, column=2).alignment = LEFT

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 22
ws1.column_dimensions['G'].width = 22
ws1.column_dimensions['H'].width = 28

# ====================================================================
# Sheet 2: 国考详情
# ====================================================================
ws2 = wb.create_sheet('国考详情')
add_title(ws2, '2026年国考温州地区岗位详细分析', 2, 1)

r = 3
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.cell(row=r, column=1, value='一、招录规模趋势')
ws2.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 4
for i, h in enumerate(['年份', '岗位数 | 招录人数 | 变化趋势'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 2)
trend = [
    ['2023年', '112个 | 256人'],
    ['2024年', '98个 | 211人 | 减少45人'],
    ['2025年', '91个 | 159人 | 减少52人'],
    ['2026年', '86个 | 138人 | 减少21人'],
]
for i, row in enumerate(trend, r+1):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    style_normal(ws2, i, 2)

r = 9
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.cell(row=r, column=1, value='二、各部门招录分布')
ws2.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 10
for i, h in enumerate(['招录部门', '人数 | 占比 | 说明'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 2)
depts = [
    ['国家税务总局浙江省税务局', '106人 | 76.8% | 绝对主力，含各县区税务分局'],
    ['浙江海事局', '12人 | 8.7% | 瓯江/乐清/飞云江海事处'],
    ['中国人民银行浙江省分行', '11人 | 8.0% | 温州市分行'],
    ['杭州海关', '6人 | 4.3% | 温州海关'],
    ['国家金融监督管理总局浙江监管局', '3人 | 2.2% | 温州监管分局'],
]
for i, row in enumerate(depts, r+1):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    style_left(ws2, i, 2)

r = 16
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.cell(row=r, column=1, value='三、学历要求分布')
ws2.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 17
for i, h in enumerate(['学历层次', '人数 | 占比'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 2)
edu = [
    ['本科及以上', '111人 | 80.4%'],
    ['仅限本科', '18人 | 13.0%'],
    ['硕士研究生及以上', '9人 | 6.5%'],
    ['本科可报合计', '129人 | 93.5%'],
]
for i, row in enumerate(edu, r+1):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    style_normal(ws2, i, 2)

r = 22
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.cell(row=r, column=1, value='四、考生身份要求')
ws2.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 23
for i, h in enumerate(['身份类型', '人数 | 岗位数 | 占比'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 2)
ident = [
    ['应届生', '90人 | 50个岗位 | 65.2%'],
    ['社会人员（不限身份）', '35人 | 26个岗位 | 25.4%'],
    ['五类人员（服务基层等）', '13人 | 10个岗位 | 9.4%'],
    ['要求党员身份', '4个岗位 | 非单独类别，叠加在其他要求上'],
]
for i, row in enumerate(ident, r+1):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    style_left(ws2, i, 2)

r = 28
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.cell(row=r, column=1, value='五、主要专业要求（税务系统为主）')
ws2.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 29
for i, h in enumerate(['专业大类', '涵盖专业举例'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 2)
majors = [
    ['经济学类', '经济学、国民经济管理、资源与环境经济学等'],
    ['财政学类', '财政学、税收学、国际税收等'],
    ['财会审计类', '会计学、财务管理、审计学等'],
    ['中国语言文学类', '汉语言文学、秘书学、应用语言学等'],
    ['新闻传播学类', '新闻学、传播学、网络与新媒体等'],
    ['法学类', '法学、知识产权等（非税务为主岗位）'],
    ['计算机类', '计算机科学与技术、软件工程等（少量岗位）'],
    ['统计学类', '统计学、应用统计学等（统计局调查队）'],
]
for i, row in enumerate(majors, r+1):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    style_left(ws2, i, 2)

r = 38
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws2.cell(row=r, column=1, value='六、关键时间节点')
ws2.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 39
for i, h in enumerate(['阶段', '时间'], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, 2)
timeline = [
    ['公告发布', '2025年10月14日'],
    ['网上报名', '2025年10月15日-10月24日'],
    ['报名确认缴费', '2025年11月初'],
    ['准考证打印', '2025年11月下旬'],
    ['笔试', '2025年11月30日'],
    ['笔试科目', '行测（9:00-11:00）+ 申论（14:00-17:00）'],
    ['成绩查询', '2026年1月'],
    ['面试', '2026年2-3月'],
]
for i, row in enumerate(timeline, r+1):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    style_normal(ws2, i, 2)

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 55

# ====================================================================
# Sheet 3: 省考详情
# ====================================================================
ws3 = wb.create_sheet('省考详情')
add_title(ws3, '2026年浙江省考温州地区岗位详细分析', 2, 1)

r = 3
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='一、招录规模变化')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 4
for i, h in enumerate(['指标', '数据'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
s1 = [
    ['2026年招录人数', '660人（495个职位）'],
    ['较2025年', '减少379人，降幅超1/3'],
    ['全省计划', '5712名（含专项预留318名）'],
    ['趋势', '连续两年下降，编制供给收紧'],
]
for i, row in enumerate(s1, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_normal(ws3, i, 2)

r = 9
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='二、岗位类别分布')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 10
for i, h in enumerate(['岗位类别', '人数 | 占比 | 说明'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
cats = [
    ['行政执法类', '309人 | 46.8% | 含人民警察95人'],
    ['综合类', '261人 | 39.5% | 含专业技术类1人'],
    ['基层类', '82人 | 12.4% | 乡镇岗位为主'],
    ['优秀村干部类', '8人 | 1.2% | 定向招录'],
]
for i, row in enumerate(cats, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_left(ws3, i, 2)

r = 15
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='三、学历与学位要求')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 16
for i, h in enumerate(['要求类型', '具体情况'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
edu2 = [
    ['本科及以上', '招录主体，已成为基本门槛'],
    ['硕士研究生及以上', '主要集中在综合类岗位'],
    ['专科可报', '极少，仅面向优秀社区干部、退役军人等定向'],
    ['要求学位', '约96%岗位要求相应学位'],
    ['学位不限', '主要集中在文成、泰顺、平阳、永嘉乡镇机关'],
]
for i, row in enumerate(edu2, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_left(ws3, i, 2)

r = 22
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='四、身份要求分布')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 23
for i, h in enumerate(['身份类型', '占比 | 说明'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
ident2 = [
    ['应届生岗位', '52.7% | 2024/2025/2026届毕业生均可报考'],
    ['不限身份', '37.0% | 社会人员可报'],
    ['其他', '约10.3% | 基层服务人员、优秀村干部等'],
]
for i, row in enumerate(ident2, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_left(ws3, i, 2)

r = 27
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='五、笔试科目分类')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 28
for i, h in enumerate(['类别', '考试科目'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
exams = [
    ['综合类', '行测A卷 + 申论A卷'],
    ['基层类', '行测B卷 + 申论B卷'],
    ['行政执法类', '行测C卷 + 申论C卷'],
    ['优秀村干部类', '行测B卷 + 综合应用能力'],
]
for i, row in enumerate(exams, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_normal(ws3, i, 2)

r = 33
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='六、关键时间节点')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 34
for i, h in enumerate(['阶段', '时间'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
tl2 = [
    ['公告发布', '2025年11月3日'],
    ['网上报名', '2025年11月6日-11月11日'],
    ['资格初审', '2025年11月6日-11月13日'],
    ['缴费确认', '2025年11月14日-11月18日'],
    ['准考证打印', '2025年12月3日-12月8日'],
    ['笔试', '2025年12月7日'],
    ['成绩查询', '2026年1月中旬'],
    ['面试', '2026年2-3月'],
]
for i, row in enumerate(tl2, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_normal(ws3, i, 2)

r = 43
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws3.cell(row=r, column=1, value='七、山区海岛县专项招录（新增）')
ws3.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 44
for i, h in enumerate(['项目', '内容'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 2)
spec = [
    ['全省计划', '324名'],
    ['温州岗位', '洞头、文成、泰顺共28名'],
    ['报考条件', '已参加2026省考笔试且成绩达原岗位最低合格线'],
    ['报名时间', '2026年4月8日-4月12日'],
    ['面试时间', '2026年4月26日'],
    ['限制', '已入围等额体检或考察的人员不得报考'],
]
for i, row in enumerate(spec, r+1):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    style_left(ws3, i, 2)

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 55

# ====================================================================
# Sheet 4: 事业单位详情
# ====================================================================
ws4 = wb.create_sheet('事业单位详情')
add_title(ws4, '2026年温州事业单位统一招聘详情', 5, 1)

headers4 = ['区县', '招聘人数', '报名时间', '笔试时间', '笔试科目', '报名平台', '应届生范围', '咨询电话']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 8)

data4 = [
    ['温州市级', '36名', '3.23-3.27', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '0577-89090125'],
    ['鹿城区', '21名', '3.17-3.24', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '鹿城区人社局'],
    ['龙湾区', '28名', '3.18-3.25', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '龙湾区人社局'],
    ['瓯海区', '28名', '3.23-3.27', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '瓯海区人社局'],
    ['洞头区', '44名', '3.19-3.25', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '洞头区人社局'],
    ['龙港市', '23名', '3.17-3.24', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '龙港市人社局'],
    ['苍南县', '41名', '3.17-3.23', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '苍南县人社局'],
    ['文成县', '63名', '3.19-3.25', '4.25', '综合应用能力+职业能力倾向测验', 'qssy.zjks.com', '2024-2026届', '文成县人社局'],
]

for i, row in enumerate(data4, 4):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)
    style_normal(ws4, i, 8)

r = 12
ws4.cell(row=r, column=1, value='合计')
ws4.cell(row=r, column=1).font = BOLD_FONT
ws4.cell(row=r, column=2, value='284名')
ws4.cell(row=r, column=2).font = RED_BOLD

r = 14
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws4.cell(row=r, column=1, value='备注：以上事业单位笔试均参加2026年4月25日全省事业单位统考')
ws4.cell(row=r, column=1).font = Font(name='微软雅黑', size=9, color=GRAY)
ws4.cell(row=r, column=1).alignment = LEFT

ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 30
ws4.column_dimensions['F'].width = 18
ws4.column_dimensions['G'].width = 16
ws4.column_dimensions['H'].width = 18

# ====================================================================
# Sheet 5: 国企详情
# ====================================================================
ws5 = wb.create_sheet('国企详情')
add_title(ws5, '2026年温州国企招聘详情', 6, 1)

headers5 = ['地区', '招聘人数', '招聘单位', '报名时间', '笔试时间', '用工性质']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)

data5 = [
    ['永嘉县', '12名', '永嘉旅游投资集团、永嘉交投集团等5家国企', '2026年3月', '2026年4月', '劳动合同制'],
    ['乐清市', '43名', '乐清市属（管）国有企业多家', '3.18-3.20', '4.11', '劳动合同制（最低3年）'],
    ['洞头区', '10名', '洞头城发建设控股集团等4家国企', '4.7-4.9', '-', '劳动合同制'],
]
for i, row in enumerate(data5, 4):
    for j, val in enumerate(row, 1):
        ws5.cell(row=i, column=j, value=val)
    style_left(ws5, i, 6)

r = 7
ws5.cell(row=r, column=1, value='合计')
ws5.cell(row=r, column=1).font = BOLD_FONT
ws5.cell(row=r, column=2, value='65名')
ws5.cell(row=r, column=2).font = RED_BOLD

ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 12
ws5.column_dimensions['C'].width = 40
ws5.column_dimensions['D'].width = 16
ws5.column_dimensions['E'].width = 12
ws5.column_dimensions['F'].width = 22

# ====================================================================
# Sheet 6: 硕博人才招引
# ====================================================================
ws6 = wb.create_sheet('硕博人才招引')
add_title(ws6, '2026年温州市硕博人才公开招引详情', 2, 1)

r = 3
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws6.cell(row=r, column=1, value='一、引进规模')
ws6.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 4
for i, h in enumerate(['类别', '单位数 | 招引人数'], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 2)
dr = [
    ['事业单位', '31家 | 213人'],
    ['知名企业', '34家 | 232人'],
    ['合计', '65家 | 445人'],
]
for i, row in enumerate(dr, r+1):
    ws6.cell(row=i, column=1, value=row[0])
    ws6.cell(row=i, column=2, value=row[1])
    style_normal(ws6, i, 2)

r = 8
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws6.cell(row=r, column=1, value='二、学历要求（符合任一即可）')
ws6.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 9
for i, h in enumerate(['条件', '具体要求'], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 2)
reqs = [
    ['条件1', '全国普通高校博士研究生'],
    ['条件2', '国（境）外高校博士研究生'],
    ['条件3', '"双一流"高校全日制硕士研究生'],
    ['条件4', '浙江省12所重点建设本科院校+5所省市共建高校的硕士'],
    ['条件5', 'ARWU/THE/QS/US.News世界前200高校硕士'],
    ['条件6', '软科医药类前30高校硕士（限卫技岗）'],
    ['条件7', '软科体育类前10高校硕士（限体育岗）'],
    ['注意', '知名企业岗位不受学校范围限制'],
]
for i, row in enumerate(reqs, r+1):
    ws6.cell(row=i, column=1, value=row[0])
    ws6.cell(row=i, column=2, value=row[1])
    style_left(ws6, i, 2)

r = 18
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws6.cell(row=r, column=1, value='三、时间安排')
ws6.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 19
for i, h in enumerate(['阶段', '时间'], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 2)
tl3 = [
    ['公告发布', '2026年4月10日'],
    ['报名时间', '4月13日9:00-4月28日17:00'],
    ['资格审核完成', '4月30日前'],
    ['国内硕士取得学位截止', '2026年7月31日'],
    ['博士取得学位截止', '2026年12月31日'],
    ['国外高校取得认证截止', '2027年1月31日'],
]
for i, row in enumerate(tl3, r+1):
    ws6.cell(row=i, column=1, value=row[0])
    ws6.cell(row=i, column=2, value=row[1])
    style_normal(ws6, i, 2)

r = 26
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws6.cell(row=r, column=1, value='四、年龄要求')
ws6.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 27
for i, h in enumerate(['学历', '年龄要求'], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 2)
ages = [
    ['硕士研究生', '1987年4月1日以后出生（39岁以下）'],
    ['博士研究生', '1982年4月1日以后出生（44岁以下）'],
]
for i, row in enumerate(ages, r+1):
    ws6.cell(row=i, column=1, value=row[0])
    ws6.cell(row=i, column=2, value=row[1])
    style_normal(ws6, i, 2)

r = 30
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws6.cell(row=r, column=1, value='五、报名与咨询')
ws6.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=11, color=DARK_BLUE)
r = 31
for i, h in enumerate(['项目', '内容'], 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, 2)
info = [
    ['报名网站', '温州市人社局官网 hrss.wenzhou.gov.cn'],
    ['报名限制', '每位应聘人员限报1个岗位'],
    ['招引工作咨询', '0577-89090125 / 89090123'],
    ['网络报名技术', '0577-88853190 / 13616614180'],
]
for i, row in enumerate(info, r+1):
    ws6.cell(row=i, column=1, value=row[0])
    ws6.cell(row=i, column=2, value=row[1])
    style_left(ws6, i, 2)

ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 55

# ====================================================================
# Sheet 7: 报考建议
# ====================================================================
ws7 = wb.create_sheet('报考建议')
add_title(ws7, '2027年备考建议与报考策略', 2, 1)

r = 3
for i, h in enumerate(['项目', '建议内容'], 1):
    ws7.cell(row=r, column=i, value=h)
style_header(ws7, r, 2)

tips = [
    ['目标定位', '关注国考（税务为主）、省考（行政执法/综合类）、事业单位统考三条线'],
    ['应届生优势', '国考65%+省考53%岗位限应届，务必把握应届身份（毕业3年内均算）'],
    ['专业匹配', '经济学/财政/财会/中文/法学是国考省考热门专业；理工科关注事业单位专业岗'],
    ['山区海岛专项', '省考笔试后额外机会，洞头/文成/泰顺28名，达线即可报名'],
    ['硕博通道', '如读研，温州硕博引才445人（事业编213+企业232），竞争远小于国省考'],
    ['备考重点', '行测+申论是基础，事业编另需综合应用能力+职业能力倾向测验'],
    ['时间规划', '国考10月报名11月笔试 → 省考11月报名12月笔试 → 事业编3-4月 → 专项4月'],
    ['报名平台', '国考: scs.gov.cn | 省考: gwy.zjks.com | 事业编: qssy.zjks.com | 硕博: hrss.wenzhou.gov.cn'],
    ['编制趋势', '国考省考连续缩减，事业编相对稳定，建议多线并行不放弃任何机会'],
    ['体检提醒', '公务员体检标准严格，提前自查视力、血压、肝功能等'],
]
for i, row in enumerate(tips, r+1):
    ws7.cell(row=i, column=1, value=row[0])
    ws7.cell(row=i, column=2, value=row[1])
    style_left(ws7, i, 2)

ws7.column_dimensions['A'].width = 16
ws7.column_dimensions['B'].width = 80

# Save
out = 'c:/Users/chenx/WorkBuddy/Claw/2026年温州市招聘岗位信息汇总.xlsx'
wb.save(out)
print(f'Done: {out}')
